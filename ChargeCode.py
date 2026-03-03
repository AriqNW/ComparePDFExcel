#!/usr/bin/env python
# coding: utf-8

# In[3]:


import os
import pandas as pd
import pdfplumber
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# 1️⃣ READ EXCEL
# =========================

excel_folder = "Excel"
excel_files = os.listdir(excel_folder)

if len(excel_files) != 1:
    raise Exception("Excel folder must contain exactly one file.")

excel_path = os.path.join(excel_folder, excel_files[0])

df_original = pd.read_excel(excel_path)
df_original = df_original.iloc[5:].reset_index(drop=True)

# =========================
# 🔎 AUTO DETECT DESCRIPTION COLUMNS
# =========================

ba_column = next(
    (col for col in df_original.columns if col.endswith("DESCRIPTION_BA")),
    None
)

en_column = next(
    (col for col in df_original.columns if col.endswith("DESCRIPTION_EN")),
    None
)

results = []

# =========================
# 2️⃣ PROCESS FUNCTION
# =========================

def process_pdf_folder(pdf_folder, column_name):

    df = df_original.copy()
    folder_results = {}

    pdf_files = [
        f for f in os.listdir(pdf_folder)
        if re.match(r"^CC\d*\.pdf$", f)
    ]

    def extract_number(filename):
        match = re.search(r"\d+", filename)
        return int(match.group()) if match else 0

    pdf_files = sorted(pdf_files, key=extract_number)

    for filename in pdf_files:

        if df.empty:
            break

        file_path = os.path.join(pdf_folder, filename)

        with pdfplumber.open(file_path) as pdf:
            page = pdf.pages[0]
            left_part = page.crop((0, 300, 330, page.height))
            text = left_part.extract_text()

        if not text:
            continue

        start_keyword = "Biaya Bulanan"
        end_keyword = "Pemakaian" if "Pemakaian" in text else "TOTAL"

        start_index = text.find(start_keyword)
        end_index = text.find(end_keyword)

        if start_index == -1 or end_index == -1:
            continue

        text = text[start_index:end_index]

        countRp = text.count("Rp.")
        pdf_items = []

        for line in text.split("\n"):
            line = line.strip()
            if line in ["Biaya Bulanan", "Lain-Lain"]:
                continue

            split_line = re.split(r"Rp\.", line)
            if split_line and split_line[0].strip():
                pdf_items.append(split_line[0].strip())

        excel_items = (
            df[column_name]
            .dropna()
            .astype(str)
            .head(countRp)
            .tolist()
        )

        pdf_counter = Counter(pdf_items)
        comparing_results = []

        for item in excel_items:
            if pdf_counter[item] > 0:
                comparing_results.append("MATCH")
                pdf_counter[item] -= 1
            else:
                comparing_results.append("NOT MATCH")

        folder_results[filename] = {
            "items": excel_items,
            "results": comparing_results
        }

        df = df.iloc[countRp:].reset_index(drop=True)

    return folder_results

# =========================
# 3️⃣ PROCESS BA & EN
# =========================

ba_results = process_pdf_folder("PDF_BA", ba_column)
en_results = process_pdf_folder("PDF_EN", en_column)

# Flatten results into sequential lists
ba_compare_all = []
en_compare_all = []

for filename in sorted(ba_results.keys(), key=lambda x: int(re.search(r"\d+", x).group()) if re.search(r"\d+", x) else 0):
    ba_compare_all.extend(ba_results[filename]["results"])

for filename in sorted(en_results.keys(), key=lambda x: int(re.search(r"\d+", x).group()) if re.search(r"\d+", x) else 0):
    en_compare_all.extend(en_results[filename]["results"])

# Build result directly from original Excel
for i in range(len(df_original)):

    ba_result = ba_compare_all[i] if i < len(ba_compare_all) else "N/A"
    en_result = en_compare_all[i] if i < len(en_compare_all) else "N/A"

    results.append({
        "TITLE": df_original.loc[i, "Title"],
        "CHARGE_TYPE": df_original.loc[i, "CHARGE_TYPE"],
        "CHARGE_CODE": df_original.loc[i, "CHARGE_CODE"],
        "DISCOUNT_DESCRIPTION_BA": df_original.loc[i, "DISCOUNT_DESCRIPTION_BA"],
        "COMPARING_RESULT_BA": ba_result,
        "DISCOUNT_DESCRIPTION_EN": df_original.loc[i, "DISCOUNT_DESCRIPTION_EN"],
        "COMPARING_RESULT_EN": en_result
    })

# =========================
# 4️⃣ SAVE FILE
# =========================

result_folder = "Result"
os.makedirs(result_folder, exist_ok=True)

output_path = os.path.join(result_folder, "Final_Result.xlsx")

output_df = pd.DataFrame(results)
output_df.to_excel(output_path, index=False)

# =========================
# 5️⃣ FORMAT + COLOR + SUMMARY
# =========================

wb = load_workbook(output_path)
ws = wb.active

light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ✅ FULL BORDER FOR MAIN TABLE
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border

# Header bold + center + light blue
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = light_blue

# Counters
ba_match = ba_not = ba_na = 0
en_match = en_not = en_na = 0

for row in range(2, ws.max_row + 1):

    ba_cell = ws[f"E{row}"]
    en_cell = ws[f"G{row}"]

    # BA
    if ba_cell.value == "MATCH":
        ba_cell.fill = green
        ba_cell.value = ""
        ba_match += 1
    elif ba_cell.value == "NOT MATCH":
        ba_cell.fill = red
        ba_cell.value = ""
        ba_not += 1
    else:
        ba_cell.fill = yellow
        ba_cell.value = ""
        ba_na += 1

    ba_cell.alignment = Alignment(horizontal="center")

    # EN
    if en_cell.value == "MATCH":
        en_cell.fill = green
        en_cell.value = ""
        en_match += 1
    elif en_cell.value == "NOT MATCH":
        en_cell.fill = red
        en_cell.value = ""
        en_not += 1
    else:
        en_cell.fill = yellow
        en_cell.value = ""
        en_na += 1

    en_cell.alignment = Alignment(horizontal="center")

# =========================
# SUMMARY TABLE
# =========================

summary_col = 9  # I column (1 column gap after G)
start_row = 1

headers = ["SUMMARY", "BA", "EN"]

for i, header in enumerate(headers):
    cell = ws.cell(row=start_row, column=summary_col + i, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    cell.fill = light_blue

summary_data = [
    ("MATCH", ba_match, en_match),
    ("NOT MATCH", ba_not, en_not),
    ("N/A", ba_na, en_na),
]

for i, (label, ba_val, en_val) in enumerate(summary_data):
    r = start_row + i + 1

    ws.cell(row=r, column=summary_col, value=label)
    ws.cell(row=r, column=summary_col+1, value=ba_val)
    ws.cell(row=r, column=summary_col+2, value=en_val)

    for c in range(summary_col, summary_col+3):
        ws.cell(row=r, column=c).border = thin_border

    ws.cell(row=r, column=summary_col).alignment = Alignment(horizontal="left")
    ws.cell(row=r, column=summary_col+1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=summary_col+2).alignment = Alignment(horizontal="center")

    # ✅ Add color to summary label row
    if label == "MATCH":
        ws.cell(row=r, column=summary_col).fill = green
    elif label == "NOT MATCH":
        ws.cell(row=r, column=summary_col).fill = red
    else:
        ws.cell(row=r, column=summary_col).fill = yellow

# =========================
# AUTO WIDTH
# =========================

for column_cells in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column_cells[0].column)

    for cell in column_cells:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[column_letter].width = max_length + 6

wb.save(output_path)

print("\nDone 🎉 The comparison report is ready at:", output_path)


# In[ ]:




